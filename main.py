from datetime import datetime

from slackbot.bot import Bot
from slackbot.bot import respond_to
from slackbot.bot import listen_to
import openpyxl
from xl import SendDataToSeed

bot = Bot()


# @respond_to('練習')
# def sample(message):
#     message.send("練習です。")
#
#
# # reply メンションつきで返す
@listen_to('練習')
def sample1(message):
    message.reply("練習ですか?")
#
#
# @listen_to('出')
# def sample2(message):
#     message.send("記録しました")
#
#
# @listen_to('出勤')
# def punch_in(message):
#     print('出勤時刻登録します')
#     timestamp = datetime.now()
#     wb = openpyxl.load_workbook('勤怠管理.xlsx')
#     ws = wb.worksheets[0]
#     max_row = ws.max_row
#     ws[max_row + 1][0].value = timestamp.strftime("%Y/%m/%d")
#     ws[max_row + 1][1].value = timestamp.strftime("%H:%M")
#     wb.save('勤怠管理.xlsx')
#     message.send(f'出勤時刻登録完了しました:{timestamp.strftime("%H:%M")}')
#
#
# @listen_to('退勤')
# def punch_out(message):
#     print('退勤時刻登録します')
#     timestamp = datetime.now()
#     wb = openpyxl.load_workbook('勤怠管理.xlsx')
#     ws = wb.worksheets[0]
#     max_row = ws.max_row
#     ws[max_row][2].value = timestamp.strftime("%H:%M")
#     wb.save('勤怠管理.xlsx')
#     message.send(f'退勤時刻登録しました:{timestamp.strftime("%H:%M")}')

bot.run()


if __name__ == '__main__':
    print(1)
    # xlファイルを作成
    # x = SendDataToSeed(datetime.now(), datetime.now(), datetime.now())
    # x.sendDate()
