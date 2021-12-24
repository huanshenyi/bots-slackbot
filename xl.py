import openpyxl
import os
from datetime import datetime
# xlファイルを作成
# wb = openpyxl.Workbook()
# ws = wb.worksheets[0]
# rng = ws["A1:C1"][0]
# rng[0].value = '日付'
# rng[1].value = '出勤時刻'
# rng[2].value = '退勤時刻'
#
# # 最終行
# max_row = ws.max_row
# # 日付
# timestamp = datetime.now()
# ws[max_row + 1][0].value = timestamp.strftime("%Y/%m/%d")
# # 出勤時刻
# ws[max_row + 1][1].value = timestamp.strftime("%H:%M")
# # 退勤時刻
# ws[max_row + 1][2].value = timestamp.strftime("%H:%M")
# wb.save('勤怠管理.xlsx')


class SendDataToSeed:
    def __init__(self, dayDate=None, inDate=None, outDate=None, fileName=None):
        self.fileName = '勤怠管理.xlsx'
        self.dayDate = dayDate
        self.inDate = inDate
        self.outDate = outDate
        self.ws = self.checkFile()

    def checkFile(self) -> object:
        if not os.path.exists(self.fileName):
            wb = openpyxl.Workbook()
            ws = wb.worksheets[0]
            rng = ws["A1:C1"][0]
            rng[0].value = '日付'
            rng[1].value = '出勤時刻'
            rng[2].value = '退勤時刻'
            wb.save('勤怠管理.xlsx')
            return wb
        else:
            return openpyxl.load_workbook(self.fileName)

    def sendDate(self):
        # 最終行
        max_row = self.ws.worksheets[0].max_row
        # 日付
        timestamp = datetime.now()
        self.ws.worksheets[0][max_row + 1][0].value = timestamp.strftime("%Y/%m/%d")
        # 出勤時刻
        self.ws.worksheets[0][max_row + 1][1].value = self.inDate.strftime("%H:%M")
        # 退勤時刻
        self.ws.worksheets[0][max_row + 1][2].value = self.outDate.strftime("%H:%M")
        self.ws.save('勤怠管理.xlsx')

if __name__ == "__main__":
    s = SendDataToSeed()
    s.checkFile()