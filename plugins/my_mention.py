from datetime import datetime

from slackbot.bot import respond_to
import openpyxl


@respond_to('出勤')
def punch_in(message):
    print('出勤時刻登録します')
    timestamp = datetime.now()
    wb = openpyxl.load_workbook('勤怠管理.xlsx')
    ws = wb.worksheets[0]
    max_row = ws.max_row
    ws[max_row + 1][0].value = timestamp.strftime("%Y/%m/%d")
    ws[max_row + 1][1].value = timestamp.strftime("%H:%M")
    wb.save('勤怠管理.xlsx')
    message.send(f'出勤時刻登録完了しました:{timestamp.strftime("%H:%M")}')


@respond_to('退勤')
def punch_out(message):
    print('退勤時刻登録します')
    timestamp = datetime.now()
    wb = openpyxl.load_workbook('勤怠管理.xlsx')
    ws = wb.worksheets[0]
    max_row = ws.max_row
    ws[max_row][2].value = timestamp.strftime("%H:%M")
    wb.save('勤怠管理.xlsx')
    message.send(f'退勤時刻登録しました:{timestamp.strftime("%H:%M")}')
